import streamlit as st
import pandas as pd
import folium
from streamlit_folium import st_folium
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def to_excel(df_dict):
    """Exporta un diccionario de DataFrames a un archivo Excel en memoria con estilos."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Aplicar estilos al libro de trabajo
        workbook = writer.book
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]: # Primera fila (headers)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.border = thin_border

    output.seek(0)
    return output

def render_map(paradas_df, resultados):
    """Renderiza un mapa más atractivo con íconos y tooltips."""
    map_center = [paradas_df['lat'].mean(), paradas_df['lon'].mean()]
    m = folium.Map(location=map_center, zoom_start=12, tiles="cartodbpositron")

    # Íconos para el depósito y las paradas
    depot_icon = folium.Icon(color='red', icon='warehouse', prefix='fa')
    stop_icon = folium.Icon(color='blue', icon='circle-dot', prefix='fa')

    for _, parada in paradas_df.iterrows():
        tooltip_text = f"<b>{parada['id']}</b><br>Demanda: {parada['demanda']}"
        icon_to_use = depot_icon if parada['is_depot'] else stop_icon
        folium.Marker(
            [parada['lat'], parada['lon']], 
            tooltip=tooltip_text, 
            icon=icon_to_use
        ).add_to(m)

    if resultados:
        # Paleta de colores profesional
        colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
        depot = paradas_df[paradas_df['is_depot']].iloc[0]
        paradas_dict = {p['id']: p for _, p in paradas_df.iterrows()}
        
        for i, ruta in enumerate(resultados):
            points = [[depot['lat'], depot['lon']]] + [[paradas_dict[pid]['lat'], paradas_dict[pid]['lon']] for pid in ruta['secuencia_paradas_ids']] + [[depot['lat'], depot['lon']]]
            folium.PolyLine(
                points, 
                color=colors[i % len(colors)], 
                weight=4, 
                opacity=0.8, 
                tooltip=f"{ruta['vehiculo_id']} ({ruta['distancia_km']:.1f} km)"
            ).add_to(m)

    st_folium(m, width='100%', height=600, returned_objects=[])

def render_results_section(resultados, paradas_df):
    """Muestra la sección de resultados (métricas y tablas) en la pestaña."""
    resumen_df = pd.DataFrame(resultados)

    st.subheader("Métricas Clave de la Operación")
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Vehículos Utilizados", f"{len(resumen_df)}")
    col2.metric("Distancia Total", f"{resumen_df['distancia_km'].sum():.2f} km")
    col3.metric("Costo Total", f"${resumen_df['costo_estimado'].sum():,.2f}")
    
    st.markdown("---")
    st.subheader("Resumen por Vehículo")
    
    # Preparar DataFrame para mostrar
    df_display = resumen_df[['vehiculo_id', 'total_demanda', 'capacidad_utilizada_pct', 'distancia_km', 'costo_estimado']].copy()
    df_display.rename(columns={
        'vehiculo_id': 'Vehículo',
        'total_demanda': 'Demanda Total',
        'capacidad_utilizada_pct': '% Capacidad',
        'distancia_km': 'Distancia (km)',
        'costo_estimado': 'Costo ($)'
    }, inplace=True)
    
    st.dataframe(
        df_display,
        hide_index=True,
        use_container_width=True,
        column_config={
            "% Capacidad": st.column_config.ProgressColumn(
                "Capacidad Utilizada",
                format="%.1f%%",
                min_value=0,
                max_value=100,
            ),
            "Distancia (km)": st.column_config.NumberColumn(format="%.1f"),
            "Costo ($)": st.column_config.NumberColumn(format="$%.2f"),
        }
    )

    st.markdown("---")
    st.subheader("Descargar Informe Detallado")
    
    # Preparar datos para el informe
    paradas_dict = {p['id']: p for _, p in paradas_df.iterrows()}
    informe_sheets = {"Resumen": df_display}
    
    for _, ruta in resumen_df.iterrows():
        sheet_name = f"Ruta {ruta['vehiculo_id']}"
        # Asegurarse de que el nombre de la hoja no sea demasiado largo
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        ruta_df = pd.DataFrame([paradas_dict[pid] for pid in ruta['secuencia_paradas_ids']])
        informe_sheets[sheet_name] = ruta_df[['id', 'lat', 'lon', 'demanda']]

    excel_data = to_excel(informe_sheets)
    
    st.download_button(
        label="📥 Descargar Informe (Excel)",
        data=excel_data,
        file_name="informe_de_rutas.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
